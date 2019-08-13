import os
import openpyxl
import statistics
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.styles import Border, Font, Side


# returns a tuple with start row in 0 index, end row in 1
def find_start_end(sheet, year):
    found = False
    start_row = 0
    end_row = 0
    for col in sheet.iter_cols(max_col=1, values_only=True):
        for val in col:
            try:
                end_row += 1
                if not found:   # only adjust start if row not found
                    start_row += 1
                if val.strftime("%Y") == str(year):
                    found = True
                elif found:
                    end_row -= 1
                    break
            except AttributeError:    # if we hit a string
                pass
    return start_row, end_row


# returns a dictionary containing all "data types" (i.e. PH) and
# their corresponding column number
def get_data_types(sheet, row_num):
    data_dict = {}
    curr_row = 0
    for row in sheet.iter_rows(min_row=row_num, max_row=row_num, values_only=True):
        for val in row:
            if val is not None:
                curr_row += 1
                data_dict[val.upper()] = curr_row
    return data_dict


def is_number(num):
    try:
        float(num)
        return True
    except (ValueError, TypeError):
        return False


def add_dates(start_end, write_sheet, date_range, doc_sheet, bold):
    start = start_end[0]
    write_sheet['A1'].value = 'DATE'
    write_sheet['A1'].font = bold
    write_sheet['B1'].value = 'DAY'
    write_sheet['B1'].font = bold
    for i in range(date_range):  # number of readings in a year
        curr_date = doc_sheet['A' + str(i + start)].value.strftime('%m/%d/%Y').lstrip('0').replace('/0', '/')
        write_sheet['A' + str(i + 2)].value = datetime.strptime(curr_date, '%m/%d/%Y').date()
        write_sheet['B' + str(i + 2)].value = doc_sheet['B' + str(i + start)].value


def compile_data_to_sheet(write_sheet, sheet_num, sheet_name, num_sheets, doc_wb, year, data_types):
    bold = Font(bold=True)
    dt_keys = list(data_types.keys())
    doc_sheet = doc_wb[sheet_name + str(sheet_num)]  # i.e. KARP SYSTEM1
    start_end = find_start_end(doc_sheet, year)
    date_range = start_end[1] - start_end[0] + 1
    if sheet_num == 1:  # add the date column the first time around
        add_dates(start_end, write_sheet, date_range, doc_sheet, bold)
    # add all data into the sheet
    for data_key in dt_keys[2:]:  # add all types of data to row 1
        col_num = data_types[data_key] - 3  # -3 because of date and day columns, i.e. 1st data col will = 0
        c = write_sheet.cell(row=1, column=3 + col_num * (num_sheets + 2))  # write the column headers
        c.value = data_key
        c.font = Font(bold=True)
        if col_num != 0:  # so we don't have a random border
            c.border = Border(left=Side(border_style='medium'))
        mean_hc = write_sheet.cell(row=1, column=3 + num_sheets + col_num * (num_sheets + 2))
        mean_hc.value = 'MEAN'
        mean_hc.font = bold
        avg_mean_hc = write_sheet.cell(row=1, column=4 + num_sheets + col_num * (num_sheets + 2))
        avg_mean_hc.value = 'AVG. MEAN'
        avg_mean_hc.font = bold
        for i in range(date_range):  # copy over data
            write_sheet.cell(row=i + 2, column=(num_sheets + 2) * col_num + sheet_num + 2).value = \
                doc_sheet.cell(row=start_end[0] + i, column=data_types[data_key]).value
    return date_range


def add_means(row_num, data_types, num_sheets, write_sheet, dt_keys):
    for data_key in dt_keys[2:]:
        if data_key.upper() != 'COMMENTS' and data_key.upper() != 'COMMENT':
            col_num = data_types[data_key] - 3
            curr_col = col_num * (num_sheets + 2) + 3
            mean = 0
            count = 0  # needed as some columns contain None or a string
            for row in write_sheet.iter_rows(min_row=row_num, max_row=row_num,
                                             min_col=curr_col, max_col=curr_col + num_sheets - 1,
                                             values_only=True):
                for val in row:
                    if is_number(val):
                        count += 1
                        mean += float(val)
            if count != 0:  # if count == 0, then we know the mean of these values is not legitimate
                write_sheet.cell(row=row_num, column=curr_col + num_sheets).value = round(mean / count, 6)


def add_avg_means(write_sheet, num_sheets, data_types, dt_keys, date_range, chart_sheet):
    offset = 0  # needed for charts; makes for smoother formatting
    for data_key in dt_keys[2:]:
        if data_key.upper() != 'COMMENTS' and data_key.upper() != 'COMMENT':
            col_num = data_types[data_key] - 3
            avg_col = col_num * (num_sheets + 2) + num_sheets + 4  # this is the column which will have average of means
            mean_col = avg_col - 1  # this is the column with means in it
            avg_mean = 0
            all_means = []
            count = 0
            for repeat in range(2):  # very similar code ran twice so we can just change it slightly using conditionals
                for row_num in range(2, date_range+2):
                    if not repeat:  # still looking for average means
                        curr_mean = write_sheet.cell(row=row_num, column=mean_col).value
                        if is_number(curr_mean):
                            count += 1
                            avg_mean += curr_mean
                            all_means.append(curr_mean)
                    else:  # now actually populate the column with average means
                        curr_cell = write_sheet.cell(row=row_num, column=avg_col)
                        if count != 0:
                            curr_cell.value = round(avg_mean/count, 6)
                        else:
                            curr_cell.value = None
                        curr_cell.border = Border(right=Side(border_style='medium'))
            if all_means:  # if our array isn't empty
                write_sheet.cell(row=date_range+2, column=mean_col).value = 'STDEV:'
                write_sheet.cell(row=date_range+2, column=mean_col).font = Font(bold=True)
                write_sheet.cell(row=date_range+2, column=avg_col).value = statistics.pstdev(all_means)

            offset = add_chart(chart_sheet, data_key, mean_col, avg_col, write_sheet, date_range, data_types,
                               col_num % 2, offset)


def add_chart(chart_sheet, data_key, mean_col, avg_col, write_sheet, date_range, data_types, col_choice, offset):
    c = LineChart()
    c.display_blanks = 'span'
    c.title = data_key.upper()
    c.style = 2
    c.x_axis.title = 'Date'
    c.x_axis = DateAxis(crossAx=100)
    c.y_axis.crossAx = 500
    c.x_axis.number_format = 'm/d/y'
    c.x_axis.majorTimeUnit = 'months'
    avg_mean_data = Reference(write_sheet, min_col=avg_col, max_col=avg_col, min_row=1, max_row=date_range + 1)
    mean_data = Reference(write_sheet, min_col=mean_col, max_col=mean_col, min_row=1, max_row=date_range + 1)
    c.add_data(avg_mean_data, titles_from_data=True)
    c.add_data(mean_data, titles_from_data=True)
    dates = Reference(write_sheet, min_col=1, max_col=1, min_row=2, max_row=date_range + 1)
    c.set_categories(dates)

    # Add to either col A or col J
    if not col_choice:
        chart_sheet.add_chart(c, 'A' + str(15 * (data_types[data_key] - 3) + 1 - offset*14))
    else:
        chart_sheet.add_chart(c, 'J' + str(15 * (data_types[data_key] - 4) + 1 - offset*14))
        offset += 1
    return offset


def transfer_data(doc_wb, write_wb, sheet_name, num_sheets, wb_dir, year):
    os.chdir(wb_dir)
    write_wb.remove(write_wb['Sheet'])
    write_sheet = write_wb.create_sheet(sheet_name + ' ' + str(year))
    write_sheet.column_dimensions['A'].width = 10  # must adjust column width so date will appear
    chart_sheet = write_wb.create_sheet(sheet_name + ' CHARTS ' + str(year))
    data_types = get_data_types(doc_wb[sheet_name + str(1)], 4)
    dt_keys = list(data_types.keys())
    date_range = -1

    # compile all data to the sheet and set date_range
    for sheet_num in range(1, num_sheets+1):
        date_range = compile_data_to_sheet(write_sheet, sheet_num, sheet_name, num_sheets, doc_wb, year, data_types)

    # now to find the mean of all readings
    for row_num in range(2, date_range+2):  # get the proper row numbers
        add_means(row_num, data_types, num_sheets, write_sheet, dt_keys)

    # finding/adding the average mean
    add_avg_means(write_sheet, num_sheets, data_types, dt_keys, date_range, chart_sheet)

    write_wb.save('writebook.xlsx')


def find_similar_sheet_names(sheet_names):
    possible_names = {}
    # essentially find all sheets that end in a number and have the same "title" format
    for sheet_index in range(len(sheet_names)-1):
        if is_number(sheet_names[sheet_index][-1]):
            curr_title = sheet_names[sheet_index][:-1]
            try:  # if we've seen a title like this, increase the count by 1
                possible_names[curr_title] += 1
            except KeyError:  # otherwise, add a new key
                possible_names[curr_title] = 1
    return possible_names


def run():
    root = tk.Tk()
    frame = tk.Frame(root, width=250, height=180)
    frame.pack()
    message = tk.Label(root, height=0, text='Please select the file you want to be compiled,\n'
                                            'then hit "Exit" when you\'re done.')
    message.place(relx=0, rely=0)
    file_path = tk.StringVar()
    file_button = tk.Button(root, text='Select File', width=20,
                            command=(lambda: file_path.set(filedialog.askopenfilename())))
    file_button.place(relx=0.2, rely=0.3)
    next_button = tk.Button(root, text='Next', width=20, command=lambda: root.destroy())
    next_button.place(relx=0.2, rely=0.6)
    root.mainloop()

    file_path = file_path.get()
    file_path_split = file_path.split('/')
    doc_title = file_path_split.pop(len(file_path_split)-1)  # get title and shorten to folder directory
    file_path = '\\'.join(file_path_split)
    os.chdir(file_path)
    doc1 = openpyxl.load_workbook(doc_title)  # now actually load the workbook

    # give the user a list of sheet names to choose from
    choice_root = tk.Tk()
    choice_text = tk.Label(choice_root, text='Click on a choice below, click "Select", then click "Next".\n\n'
                                             'NOTE:  If expected system is not found on this sheet, make sure you have '
                                             '\nit listed in proper format, i.e. "KARP SYS1" through "KARP SYS6".\n')
    choice_text.pack()
    choice_list = tk.Listbox(choice_root)
    poss_sheet_names = find_similar_sheet_names(doc1.sheetnames)
    count = 1
    for sheet_name in poss_sheet_names.keys():
        choice_list.insert(count, sheet_name)
        count += 1
    choice_list.pack()
    choice_frame = tk.Frame(choice_root, width=175)
    choice_frame.pack()
    choice = tk.IntVar()
    newline = tk.Label(choice_root, text='')
    newline.pack()
    select = tk.Button(choice_root, text='Select', width=10,
                       command=lambda: choice.set(choice_list.curselection()[0]))
    select.pack()
    next_button = tk.Button(choice_root, text='Next', width=10,
                            command=lambda: choice_root.destroy())
    next_button.pack()
    choice_root.mainloop()

    # give the user a year selection spinbox
    year_root = tk.Tk()
    year_label = tk.Label(year_root, text='Please input a year below.\nNOTE: Year must be in range; '
                                          'otherwise, error will occur. \n')
    year_label.pack()
    now = datetime.now()
    year_sb = tk.Spinbox(year_root, from_=int(now.year)-1, to=5000)
    year_sb.pack()
    newline = tk.Label(year_root, text='')
    newline.pack()
    year = tk.IntVar()
    year_select = tk.Button(year_root, text='Select', width=13,
                            command=lambda: year.set(year_sb.get()))
    year_select.pack()
    compile_button = tk.Button(year_root, text='Compile', width=13,
                               command=lambda: year_root.destroy())
    compile_button.pack()
    year_root.mainloop()

    # run the compilation function
    choice = choice.get()
    sheet_title = list(poss_sheet_names.keys())[choice]
    write_book = openpyxl.Workbook()
    year = year.get()
    transfer_data(doc1, write_book, sheet_title, poss_sheet_names[sheet_title], file_path, year)

    # print a final message to let the user know the program has finished
    final = tk.Tk()
    final_text = tk.Label(final, text='Done!\n\nPlease check in the same directory as the original sheet for the final '
                          'product, denoted "writebook.xlsx".\nFeel free to edit the document as necessary.')
    final_text.pack()
    note_text = tk.Label(final, text='\nOriginal directory: ' + file_path + '\n')
    note_text.pack()
    exit_button = tk.Button(final, text='Exit', width=10,
                            command=lambda: final.destroy())
    exit_button.pack()
    final.mainloop()


run()
